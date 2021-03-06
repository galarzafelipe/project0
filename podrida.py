#Podrida#
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, FORMULAE, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

#Cuantos jugadores?
#En que orden estan sentados?
#Sube de a 1 hasta el max y baja de a 1
##############

players = input("Cuantos jugadores?")
gamemode = input("Que tipo de juego? Opciones: 1.Pares y Impares 2.Impares y Pares 3.Juego completo")
st = input("Se termina con una mano con triunfo?(Si o No)")
players = int(players)
gamemode = int(gamemode)
maxcards = int(52 / players)
reductor = maxcards + 1
players = int(players)
playerList = []
evens = []
odd = []
numseparator = list(range(1, maxcards+1))
for num in numseparator:
    if num % 2 == 0:
        evens.append(num)
    else:
        odd.append(num)
nums = numseparator[0:-1]
nums = nums[::-1]
reven = evens[::-1]
rodd = odd[::-1]
if st == 'Si':
    mxc = str(maxcards)
    y = mxc + " CT"
    reven.append(y)
    rodd.append(y)
    nums.append(y)
    print(reven)
    print(rodd)
    print(nums)
lastcol = 3 * players + 2
for i in range(players):
    p = input("Please enter Player {}:".format(i + 1))
    playerList.append(p)

workbook = Workbook()
sheet = workbook.active
align = Alignment(horizontal = 'left', vertical = 'center')
pedidas = []
hechas = []
puntos = []

#Preparing Workbook#
if gamemode == 3:
    fhalf = int(len(numseparator))
    lhalf = int(len(nums))
    lastrow = fhalf + lhalf + players + 2
    for x in range(len(numseparator)):
        sheet.cell(row = x + 2, column = 1).value = numseparator[x]
        sheet.cell(row = x + 2, column = 1).alignment = align
    for y in range(len(nums)):
        sheet.cell(row = fhalf + players + y + 2, column = 1).value = nums[y]
        sheet.cell(row = fhalf + players + y + 2, column = 1).alignment = align
    for y in range(1,players + 1,1):
        sheet.cell(row = 1, column = 3*y-1).value = playerList[y - 1]
        sheet.cell(row = 1, column = 3*y-1).font = Font(bold = True)
        sheet.cell(row = 1, column = 3*y).value = "P?"
        cw = get_column_letter(3*y-1)
        cw1 = get_column_letter(3*y)
        sheet.column_dimensions[cw1].width = 4
        sheet.cell(row = 1, column = 3*y+1).value = "H?"
        cw2 = get_column_letter(3*y+1)
        sheet.column_dimensions[cw2].width = 4
        sheet.cell(row = y + fhalf + 1, column = 1).value = "{} ST".format(maxcards)
        sheet.cell(row = fhalf + lhalf + players + 2, column = 1).value = "Puntos"
        sheet.cell(row = fhalf + lhalf + players + 3, column = 1).value = "Puestos"
        sheet.merge_cells(start_row=fhalf + lhalf + players + 3,start_column=3*y-1, end_row=fhalf + lhalf + players + 3, end_column = 3*y+1)
        pedidas.append(cw1)
        hechas.append(cw2)
        puntos.append(cw)

if gamemode == 1:
    fhalf = int(len(evens))
    lhalf = int(len(rodd))
    lastrow = fhalf + lhalf + players + 2
    for x in range(len(evens)):
        sheet.cell(row = x + 2, column = 1).value = evens[x]
        sheet.cell(row = x + 2, column = 1).alignment = align
    for y in range(len(rodd)):
        sheet.cell(row = fhalf + players + y + 2, column = 1).value = rodd[y]
        sheet.cell(row = fhalf + players + y + 2, column = 1).alignment = align
    for y in range(1,players + 1,1):
        sheet.cell(row = 1, column = 3*y-1).value = playerList[y - 1]
        sheet.cell(row = 1, column = 3*y-1).font = Font(bold = True)
        sheet.cell(row = 1, column = 3*y).value = "P?"
        cw = get_column_letter(3*y-1)
        cw1 = get_column_letter(3*y)
        sheet.column_dimensions[cw1].width = 4
        sheet.cell(row = 1, column = 3*y+1).value = "H?"
        cw2 = get_column_letter(3*y+1)
        sheet.column_dimensions[cw2].width = 4
        sheet.cell(row = y + fhalf + 1, column = 1).value = "{} ST".format(maxcards)
        sheet.cell(row = fhalf + lhalf + players + 2, column = 1).value = "Puntos"
        sheet.cell(row = fhalf + lhalf + players + 3, column = 1).value = "Puestos"
        sheet.merge_cells(start_row=fhalf + lhalf + players + 3,start_column=3*y-1, end_row=fhalf + lhalf + players + 3, end_column = 3*y+1)
        pedidas.append(cw1)
        hechas.append(cw2)
        puntos.append(cw)

    sheet.cell(row = 1, column = lastcol).value = "Basas Pedidas"
    sheet.cell(row = 1, column = lastcol +1).value = "Basas Completas"

if gamemode == 2:
    fhalf = int(len(odd))
    lhalf = int(len(reven))
    lastrow = fhalf + lhalf + players + 2
    for x in range(len(odd)):
        sheet.cell(row = x + 2, column = 1).value = odd[x]
        sheet.cell(row = x + 2, column = 1).alignment = align
    for y in range(len(reven)):
        sheet.cell(row = fhalf + players + y + 2, column = 1).value = reven[y]
        sheet.cell(row = fhalf + players + y + 2, column = 1).alignment = align
    for y in range(1,players + 1,1):
        sheet.cell(row = 1, column = 3*y-1).value = playerList[y - 1]
        sheet.cell(row = 1, column = 3*y-1).font = Font(bold = True)
        sheet.cell(row = 1, column = 3*y).value = "P?"
        cw = get_column_letter(3*y-1)
        cw1 = get_column_letter(3*y)
        sheet.column_dimensions[cw1].width = 4
        sheet.cell(row = 1, column = 3*y+1).value = "H?"
        cw2 = get_column_letter(3*y+1)
        sheet.column_dimensions[cw2].width = 4
        sheet.cell(row = y + fhalf + 1, column = 1).value = "{} ST".format(maxcards)
        sheet.cell(row = fhalf + lhalf + players + 2, column = 1).value = "Puntos"
        sheet.cell(row = fhalf + lhalf + players + 3, column = 1).value = "Puestos"
        sheet.merge_cells(start_row=fhalf + lhalf + players + 3,start_column=3*y-1, end_row=fhalf + lhalf + players + 3, end_column = 3*y+1)
        pedidas.append(cw1)
        hechas.append(cw2)
        puntos.append(cw)

#BORDERS#
thin_border = Border(left=Side(style='none'),
                     right=Side(style='thin'),
                     top=Side(style='none'),
                     bottom=Side(style='none'))

thick_border = Border(left=Side(style='none'),
                     right=Side(style='thick'),
                     top=Side(style='none'),
                     bottom=Side(style='none'))

thick_border1 = Border(left=Side(style='none'),
                     right=Side(style='thick'),
                     top=Side(style='thick'),
                     bottom=Side(style='thick'))

thin_border2 = Border(left=Side(style='none'),
                     right=Side(style='thin'),
                     top=Side(style='thick'),
                     bottom=Side(style='thick'))

for z in range(1,lastrow+1,1):
    for j in range(1, players + 1, 1):
            sheet.cell(row = z, column = 3 * j - 1).border = thin_border
            sheet.cell(row = z, column = 3 * j).border = thin_border
            sheet.cell(row = z, column = 3 * j + 1).border = thick_border
            sheet.cell(row = z, column = lastcol).border = thick_border1
            sheet.cell(row = z, column = lastcol + 1).border = thick_border1
    sheet.cell(row = z, column = 1).border = thick_border

for j in range(1, players + 1, 1):
    sheet.cell(row = lastrow, column = 3 * j - 1).border = thin_border2
    sheet.cell(row = lastrow, column = 3 * j).border = thin_border2
    sheet.cell(row = lastrow, column = 3 * j + 1).border = thick_border1
    sheet.cell(row = lastrow+1, column = 3 * j - 1).border = thin_border2
    sheet.cell(row = lastrow+1, column = 3 * j).border = thin_border2
    sheet.cell(row = lastrow+1, column = 3 * j + 1).border = thick_border1
sheet.cell(row = lastrow, column = 1).border = thick_border1
sheet.cell(row = lastrow+1, column = 1).border = thick_border1
#BORDERS#
##FUNCTIONALITY###
if players == 6:
    for i in range(1, lastrow, 1):
        newi = str(i)
        n1 = pedidas[0] + newi
        n2 = pedidas[1] + newi
        n3 = pedidas[2] + newi
        n4 = pedidas[3] + newi
        n5 = pedidas[4] + newi
        n6 = pedidas[5] + newi
        m1 = hechas[0] + newi
        m2 = hechas[1] + newi
        m3 = hechas[2] + newi
        m4 = hechas[3] + newi
        m5 = hechas[4] + newi
        m6 = hechas[5] + newi
        sheet.cell(row = i, column = lastcol).value = "=SUM({},{},{},{},{},{})".format(n1,n2,n3,n4,n5,n6)
        sheet.cell(row = i, column = lastcol+1).value = "=SUM({},{},{},{},{},{})".format(m1,m2,m3,m4,m5,m6)
    for i in range(len(puntos)):
        b = column_index_from_string(puntos[i])
        aa = puntos[i] + '2'
        bb = pedidas[i] + '2'
        cc = hechas[i] + '2'
        sheet.cell(row = 2, column = b).value = "=IF(ISBLANK({}),,IF({}={},10+2*{},{}))".format(cc,cc,bb,bb,cc)
        lr = str(lastrow - 1)
        aaa = puntos[i] + lr
        sheet.cell(row = lastrow, column = b).value = "=SUM({}:{})".format(aa,aaa)
        for z in range (3,lastrow, 1):
            newz = str(z)
            oldz = str(z-1)
            oldpun = puntos[i] + oldz
            pun = puntos[i] + newz
            ped = pedidas[i] + newz
            hec = hechas[i] + newz
            sheet.cell(row = z, column = b).value = "=IF(ISBLANK({}),,IF({}={},{}+10+2*{},{}+{}))".format(hec,hec,ped,oldpun,ped,oldpun,hec)
    lrz = str(lastrow)
    lrza = str(lastrow + 1)
    p1 = puntos[0] + lrz
    p2 = puntos[1] + lrz
    p3 = puntos[2] + lrz
    p4 = puntos[3] + lrz
    p5 = puntos[4] + lrz
    p6 = puntos[5] + lrz
    sheet.cell(row = lastrow+1, column = 2).value = "=RANK({},({},{},{},{},{},{}),0)".format(p1,p1,p2,p3,p4,p5,p6)
    sheet.cell(row = lastrow+1, column = 5).value = "=RANK({},({},{},{},{},{},{}),0)".format(p2,p1,p2,p3,p4,p5,p6)
    sheet.cell(row = lastrow+1, column = 8).value = "=RANK({},({},{},{},{},{},{}),0)".format(p3,p1,p2,p3,p4,p5,p6)
    sheet.cell(row = lastrow+1, column = 11).value = "=RANK({},({},{},{},{},{},{}),0)".format(p4,p1,p2,p3,p4,p5,p6)
    sheet.cell(row = lastrow+1, column = 14).value = "=RANK({},({},{},{},{},{},{}),0)".format(p5,p1,p2,p3,p4,p5,p6)
    sheet.cell(row = lastrow+1, column = 17).value = "=RANK({},({},{},{},{},{},{}),0)".format(p6,p1,p2,p3,p4,p5,p6)

if players == 5:
    for i in range(1, lastrow, 1):
        newi = str(i)
        n1 = pedidas[0] + newi
        n2 = pedidas[1] + newi
        n3 = pedidas[2] + newi
        n4 = pedidas[3] + newi
        n5 = pedidas[4] + newi
        m1 = hechas[0] + newi
        m2 = hechas[1] + newi
        m3 = hechas[2] + newi
        m4 = hechas[3] + newi
        m5 = hechas[4] + newi

        sheet.cell(row = i, column = lastcol).value = "=SUM({},{},{},{},{})".format(n1,n2,n3,n4,n5)
        sheet.cell(row = i, column = lastcol+1).value = "=SUM({},{},{},{},{})".format(m1,m2,m3,m4,m5)

    for i in range(len(puntos)):
        b = column_index_from_string(puntos[i])
        aa = puntos[i] + '2'
        bb = pedidas[i] + '2'
        cc = hechas[i] + '2'
        sheet.cell(row = 2, column = b).value = "=IF(ISBLANK({}),,IF({}={},10+2*{},{}))".format(cc,cc,bb,bb,cc)
        lr = str(lastrow - 1)
        aaa = puntos[i] + lr
        sheet.cell(row = lastrow, column = b).value = "=SUM({}:{})".format(aa,aaa)
        for z in range (3,lastrow, 1):
            newz = str(z)
            oldz = str(z-1)
            oldpun = puntos[i] + oldz
            pun = puntos[i] + newz
            ped = pedidas[i] + newz
            hec = hechas[i] + newz
            sheet.cell(row = z, column = b).value = "=IF(ISBLANK({}),,IF({}={},{}+10+2*{},{}+{}))".format(hec,hec,ped,oldpun,ped,oldpun,hec)
    lrz = str(lastrow)
    p1 = puntos[0] + lrz
    p2 = puntos[1] + lrz
    p3 = puntos[2] + lrz
    p4 = puntos[3] + lrz
    p5 = puntos[4] + lrz
    sheet.cell(row = lastrow+1, column = 2).value = "=RANK({},({},{},{},{},{}),0)".format(p1,p1,p2,p3,p4,p5)
    sheet.cell(row = lastrow+1, column = 5).value = "=RANK({},({},{},{},{},{}),0)".format(p2,p1,p2,p3,p4,p5)
    sheet.cell(row = lastrow+1, column = 8).value = "=RANK({},({},{},{},{},{}),0)".format(p3,p1,p2,p3,p4,p5)
    sheet.cell(row = lastrow+1, column = 11).value = "=RANK({},({},{},{},{},{}),0)".format(p4,p1,p2,p3,p4,p5)
    sheet.cell(row = lastrow+1, column = 14).value = "=RANK({},({},{},{},{},{}),0)".format(p5,p1,p2,p3,p4,p5)
    sheet.cell(row = lastrow+1, column = 2).alignment = align
    sheet.cell(row = lastrow+1, column = 5).alignment = align
    sheet.cell(row = lastrow+1, column = 8).alignment = align
    sheet.cell(row = lastrow+1, column = 11).alignment = align
    sheet.cell(row = lastrow+1, column = 14).alignment = align

if players == 4:
    for i in range(1, lastrow, 1):
        newi = str(i)
        n1 = pedidas[0] + newi
        n2 = pedidas[1] + newi
        n3 = pedidas[2] + newi
        n4 = pedidas[3] + newi
        m1 = hechas[0] + newi
        m2 = hechas[1] + newi
        m3 = hechas[2] + newi
        m4 = hechas[3] + newi

        sheet.cell(row = i, column = lastcol).value = "=SUM({},{},{},{})".format(n1,n2,n3,n4)
        sheet.cell(row = i, column = lastcol+1).value = "=SUM({},{},{},{})".format(m1,m2,m3,m4)
    for i in range(len(puntos)):
        b = column_index_from_string(puntos[i])
        aa = puntos[i] + '2'
        bb = pedidas[i] + '2'
        cc = hechas[i] + '2'
        sheet.cell(row = 2, column = b).value = "=IF(ISBLANK({}),,IF({}={},10+2*{},{}))".format(cc,cc,bb,bb,cc)
        lr = str(lastrow - 1)
        aaa = puntos[i] + lr
        sheet.cell(row = lastrow, column = b).value = "=SUM({}:{})".format(aa,aaa)
        for z in range (3,lastrow, 1):
            newz = str(z)
            oldz = str(z-1)
            oldpun = puntos[i] + oldz
            pun = puntos[i] + newz
            ped = pedidas[i] + newz
            hec = hechas[i] + newz
            sheet.cell(row = z, column = b).value = "=IF(ISBLANK({}),,IF({}={},{}+10+2*{},{}+{}))".format(hec,hec,ped,oldpun,ped,oldpun,hec)
        lrz = str(lastrow)
        p1 = puntos[0] + lrz
        p2 = puntos[1] + lrz
        p3 = puntos[2] + lrz
        p4 = puntos[3] + lrz
        sheet.cell(row = lastrow+1, column = 2).value = "=RANK({},({},{},{},{}),0)".format(p1,p1,p2,p3,p4)
        sheet.cell(row = lastrow+1, column = 5).value = "=RANK({},({},{},{},{}),0)".format(p2,p1,p2,p3,p4)
        sheet.cell(row = lastrow+1, column = 8).value = "=RANK({},({},{},{},{}),0)".format(p3,p1,p2,p3,p4)
        sheet.cell(row = lastrow+1, column = 11).value = "=RANK({},({},{},{},{}),0)".format(p4,p1,p2,p3,p4)
        sheet.cell(row = lastrow+1, column = 2).alignment = align
        sheet.cell(row = lastrow+1, column = 5).alignment = align
        sheet.cell(row = lastrow+1, column = 8).alignment = align
        sheet.cell(row = lastrow+1, column = 11).alignment = align

sheet.cell(row = 1, column = lastcol).value = "Basas Pedidas"
lastone = get_column_letter(lastcol)
bpsum = lastone + '2'
endsum = str(lastrow - 1)
bpsum1 = lastone + endsum
sheet.column_dimensions[lastone].width = 15
sheet.cell(row = lastrow, column = lastcol).value = "=SUM({}:{})".format(bpsum, bpsum1)
sheet.cell(row = 1, column = lastcol +1).value = "Basas Completas"
lastone1 = get_column_letter(lastcol + 1)
sheet.column_dimensions[lastone1].width = 17
bhsum = lastone1 + '2'
bhsum1 = lastone1 + endsum
sheet.cell(row = lastrow, column = lastcol+1).value = "=SUM({}:{})".format(bhsum, bhsum1)
#Preparing Workbook#
workbook.save(filename = "Trial1.xlsx")
print("Enjoy Playing!")
